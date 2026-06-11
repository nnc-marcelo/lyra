import streamlit as st
import pandas as pd
import openpyxl
import io
import re
from io import BytesIO
from decimal import Decimal
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
pd.set_option('display.max_colwidth', None)

st.title("Processamento de Relatórios")
st.caption("Transforma relatórios e prepara-os para o Reprtoir.")

template = st.selectbox("STATEMENT TEMPLATE:", ["Nikita Digital", "Backoffice", "YouTube (Consolidação)", "Warner Chappell"])

# ============================================================================
# TEMPLATE: NIKITA DIGITAL
# ============================================================================

# Para o template Nikita, geramos um arquivo por aba (por posição, base 1).
# "last" = última aba.
NIKITA_OUTPUTS = [
    ("DSP", 2),                 # segunda aba
    ("YouTube", 3),             # terceira aba
    ("YouTube Music", "last"),  # última aba
]


def build_single_sheet(raw_bytes, keep_idx):
    """Retorna um xlsx em memória contendo apenas a aba do índice informado."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    keep_name = wb.sheetnames[keep_idx]
    for name in list(wb.sheetnames):
        if name != keep_name:
            del wb[name]
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, keep_name


def render_nikita():
    uploaded = st.file_uploader("Upload do relatório (.xlsx)", type=["xlsx"])
    if not uploaded:
        return

    raw = uploaded.read()
    sheets = openpyxl.load_workbook(io.BytesIO(raw), read_only=True).sheetnames
    base_name = uploaded.name.rsplit(".", 1)[0]

    st.success(f"Arquivo carregado com {len(sheets)} aba(s). Baixe cada relatório abaixo:")

    for label, rule in NIKITA_OUTPUTS:
        keep_idx = len(sheets) - 1 if rule == "last" else rule - 1

        if keep_idx < 0 or keep_idx >= len(sheets):
            st.warning(f"{label}: aba esperada não encontrada (arquivo tem {len(sheets)} abas).")
            continue

        data, keep_name = build_single_sheet(raw, keep_idx)
        st.download_button(
            label=f"Baixar {label}  (aba: {keep_name})",
            data=data,
            file_name=f"{base_name}_{label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"nikita_{label}",
        )


# ============================================================================
# TEMPLATE: BACKOFFICE
# ============================================================================

def detectar_formato_arquivo(file):
    """
    Detecta automaticamente o formato do arquivo Backoffice.

    Retorna:
        - 0: arquivo começa direto com o cabeçalho (header=0)
        - 5: arquivo tem metadados nas primeiras linhas (header=5)
        - None: arquivo inválido ou muito pequeno
    """
    try:
        df_test = pd.read_excel(file, nrows=7)

        if len(df_test) < 1:
            return None

        if "BO_PayeesID" in df_test.columns or "PAYEESID" in str(df_test.columns[0]).upper():
            return 0

        if len(df_test) >= 6:
            file.seek(0)
            df_test_h5 = pd.read_excel(file, header=5, nrows=1)
            if "BO_PayeesID" in df_test_h5.columns or any("PAYEE" in str(col).upper() for col in df_test_h5.columns):
                return 5

        return None

    except Exception:
        return None
    finally:
        file.seek(0)


def ler_arquivo_backoffice(file):
    """Lê arquivo Backoffice detectando automaticamente o formato."""
    header_pos = detectar_formato_arquivo(file)

    if header_pos is None:
        return None, "❌ Arquivo muito pequeno ou formato inválido", False

    try:
        df = pd.read_excel(file, header=header_pos)
        if len(df) == 0:
            return None, "⚠️ Arquivo sem dados", False
        info = f"✅ Lido com sucesso (header={header_pos}, {len(df)} linhas)"
        return df, info, True
    except Exception as e:
        return None, f"❌ Erro ao ler: {str(e)}", False


def render_backoffice():
    st.caption("Concatena e totaliza os arquivos Backoffice para conferência e inclusão no Repertoir.")

    uploaded_files = st.file_uploader(
        "Faça o upload dos arquivos Excel",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="concat_files",
    )

    if not uploaded_files:
        st.info("📤 Aguardando upload dos arquivos...")
        st.markdown("""
        ### 📝 Instruções:

        **Para Concatenar:**
        - Faça upload de múltiplos arquivos Excel do Backoffice
        - Clique em "🔗 Concatenar arquivos"
        - Baixe o arquivo único com todos os dados

        **Para Totalizar:**
        - Faça upload de arquivos **ST** (Statements)
        - Clique em "🧮 Calcular totais"
        - Visualize o resumo financeiro e baixe os totais
        """)
        return

    st.info(f"📁 {len(uploaded_files)} arquivo(s) carregado(s)")

    col1, col2 = st.columns(2)
    with col1:
        concat_button = st.button('🔗 Concatenar arquivos', type='secondary', use_container_width=True)
    with col2:
        totals_button = st.button('🧮 Calcular totais', type='primary', use_container_width=True)

    # ---- CONCATENAR ----
    if concat_button:
        st.divider()
        st.subheader("📊 Processando concatenação...")

        dataframes, logs = [], []
        arquivos_sucesso = arquivos_erro = 0

        progress_bar = st.progress(0)
        status_text = st.empty()

        for i, file in enumerate(uploaded_files):
            progress_bar.progress((i + 1) / len(uploaded_files))
            status_text.text(f"Processando {i+1}/{len(uploaded_files)}: {file.name}")

            df, info, sucesso = ler_arquivo_backoffice(file)
            if sucesso:
                dataframes.append(df)
                arquivos_sucesso += 1
            else:
                arquivos_erro += 1
            logs.append(f"**{file.name}** - {info}")

        progress_bar.empty()
        status_text.empty()

        with st.expander(f"📋 Detalhes do processamento ({arquivos_sucesso} sucesso, {arquivos_erro} erro)", expanded=False):
            for log in logs:
                st.markdown(log)

        if dataframes:
            try:
                concatenated_df = pd.concat(dataframes, ignore_index=True)
                st.success(f"""
                ✅ **Concatenação concluída com sucesso!**
                - Arquivos processados: {arquivos_sucesso}/{len(uploaded_files)}
                - Total de linhas: {len(concatenated_df):,}
                - Total de colunas: {len(concatenated_df.columns)}
                """)

                with st.expander("👁️ Visualizar dados concatenados", expanded=False):
                    st.dataframe(concatenated_df.head(100), use_container_width=True)

                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    concatenated_df.to_excel(writer, index=False, sheet_name='Dados Concatenados')

                st.download_button(
                    label="📥 Baixar arquivo concatenado",
                    data=buffer.getvalue(),
                    file_name="backoffice_concatenado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"❌ Erro ao concatenar os arquivos: {str(e)}")
        else:
            st.error("❌ Nenhum arquivo pôde ser processado com sucesso!")

    # ---- CALCULAR TOTAIS ----
    if totals_button:
        st.divider()
        st.subheader("💰 Calculando totais...")

        results = []
        arquivos_processados = 0
        arquivos_ignorados = []

        progress_bar = st.progress(0)
        status_text = st.empty()

        for i, file in enumerate(uploaded_files):
            progress_bar.progress((i + 1) / len(uploaded_files))
            status_text.text(f"Processando {i+1}/{len(uploaded_files)}: {file.name}")

            if "ST" in file.name.upper():
                try:
                    df, info, sucesso = ler_arquivo_backoffice(file)
                    if not sucesso:
                        arquivos_ignorados.append((file.name, info))
                        continue

                    coluna_royalties = None
                    if "ROYALTIES_TO_BE_PAID" in df.columns:
                        coluna_royalties = "ROYALTIES_TO_BE_PAID"
                    elif "ROYALTIES_TO_BE_PAID_$" in df.columns:
                        coluna_royalties = "ROYALTIES_TO_BE_PAID_$"
                    else:
                        for col in df.columns:
                            if "ROYALTIES" in str(col).upper() and "PAID" in str(col).upper():
                                coluna_royalties = col
                                break

                    if coluna_royalties:
                        total_royalties = df[coluna_royalties].sum()
                        results.append((file.name, total_royalties))
                        arquivos_processados += 1
                    else:
                        arquivos_ignorados.append((file.name, "❌ Coluna de royalties não encontrada"))

                except Exception as e:
                    arquivos_ignorados.append((file.name, f"❌ Erro: {str(e)}"))
            else:
                arquivos_ignorados.append((file.name, "⚠️ Não é arquivo ST (Statement)"))

        progress_bar.empty()
        status_text.empty()

        if arquivos_ignorados:
            with st.expander(f"⚠️ Arquivos ignorados ({len(arquivos_ignorados)})", expanded=False):
                for nome, motivo in arquivos_ignorados:
                    st.markdown(f"**{nome}** - {motivo}")

        if results:
            df_results = pd.DataFrame(results, columns=["Arquivo", "Soma de ROYALTIES_TO_BE_PAID"])
            df_results["Soma de ROYALTIES_TO_BE_PAID"] = df_results["Soma de ROYALTIES_TO_BE_PAID"].round(2)
            total_royalties_sum = df_results["Soma de ROYALTIES_TO_BE_PAID"].sum().round(2)
            df_results.loc[len(df_results.index)] = ["TOTAL GERAL", total_royalties_sum]
            df_results["Soma de ROYALTIES_TO_BE_PAID"] = df_results["Soma de ROYALTIES_TO_BE_PAID"].apply(
                lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )

            st.success(f"✅ {arquivos_processados} arquivo(s) totalizado(s) com sucesso!")
            st.dataframe(df_results, use_container_width=True, hide_index=True)

            st.divider()
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric(
                    label="💰 Total Bruto",
                    value=f"R$ {total_royalties_sum:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                )
            desconto_r3 = (total_royalties_sum * 0.025).round(2)
            with col2:
                st.metric(
                    label="📉 Desconto R3 (2,5%)",
                    value=f"R$ {desconto_r3:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                )
            total_liquido = (total_royalties_sum - desconto_r3).round(2)
            with col3:
                st.metric(
                    label="✅ Total Líquido",
                    value=f"R$ {total_liquido:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                )

            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                df_export = pd.DataFrame(results, columns=["Arquivo", "Total_Royalties"])
                df_export.loc[len(df_export.index)] = ["TOTAL GERAL", total_royalties_sum]
                df_export.loc[len(df_export.index)] = ["Desconto R3 (2,5%)", desconto_r3]
                df_export.loc[len(df_export.index)] = ["TOTAL LÍQUIDO", total_liquido]
                df_export.to_excel(writer, index=False, sheet_name='Totais')

            st.download_button(
                label="📥 Baixar totais em Excel",
                data=buffer.getvalue(),
                file_name="totais_backoffice.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.error("❌ Nenhum arquivo válido para totalização foi encontrado.")


# ============================================================================
# TEMPLATE: YOUTUBE (CONSOLIDAÇÃO)
# ============================================================================

# Template-alvo: layout "asset raw" (27 colunas). Os demais formatos que o YouTube
# envia são remapeados para este template (detecção automática):
#   - "red label" (assinatura): começa com a linha "Asset Summary" e usa "Month";
#   - "ads/AdSense": não tem coluna de período nem Adjustment Type/Asset Type.
# Colunas sem equivalente ficam vazias; campos exigidos pelo Reprtoir são preenchidos.
YT_TARGET = [
    "Adjustment Type", "Day", "Country", "Asset ID", "Asset Title", "Asset Labels",
    "Asset Channel ID", "Asset Type", "Custom ID", "ISRC", "UPC", "GRid", "Artist",
    "Album", "Label", "Administer Publish Rights", "Owned Views",
    "YouTube Revenue Split : Auction", "YouTube Revenue Split : Reserved",
    "YouTube Revenue Split : Partner Sold YouTube Served",
    "YouTube Revenue Split : Partner Sold Partner Served", "YouTube Revenue Split",
    "Partner Revenue : Auction", "Partner Revenue : Reserved",
    "Partner Revenue : Partner Sold YouTube Served",
    "Partner Revenue : Partner Sold Partner Served", "Partner Revenue",
]


def _yt_fmt_money(x):
    """Formata em pt-BR com prefixo US$ (receita YouTube é em dólar)."""
    return f"US$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _yt_map_columns(src_cols):
    """Mapeia cada coluna do template para a coluna de origem (ou None)."""
    norm = {c.strip(): c for c in src_cols}

    def find(*names):
        for n in names:
            if n in norm:
                return norm[n]
        return None

    mapping = {}
    for t in YT_TARGET:
        if t == "Day":
            mapping[t] = find("Day", "Month")      # red_label usa "Month"
        elif t == "GRid":
            mapping[t] = find("GRid", "GRID")       # red_label usa "GRID"
        else:
            mapping[t] = find(t)
    return mapping


def _yt_normalize_day(v, fallback=""):
    """Unifica o período em AAAAMMDD. AAAAMM (red label) -> dia 01;
    vazio (ads, sem período) -> data do statement (extraída do nome do arquivo)."""
    s = str(v).strip()
    if s == "":
        return fallback
    if len(s) == 6 and s.isdigit():
        return s + "01"
    return s


def _yt_stmt_date(name):
    """Extrai a data AAAAMMDD do nome do arquivo (usada quando o relatório não traz período)."""
    found = re.findall(r"\d{8}", name)
    return found[-1] if found else ""


def _yt_read_one(name, raw):
    """Lê um relatório YouTube (.csv) e devolve o DataFrame já no template asset raw.
    Aceita os formatos asset raw, red label e ads/AdSense (detecção automática)."""
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    lines = text.splitlines()
    hdr_i = None
    for i, l in enumerate(lines):
        if "Country" in l and "Partner Revenue" in l:
            hdr_i = i
            break
    if hdr_i is None:
        return None, "❌ cabeçalho (Country/Partner Revenue) não encontrado"

    csv_text = "\n".join(lines[hdr_i:])
    df = pd.read_csv(io.StringIO(csv_text), dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    mapping = _yt_map_columns(df.columns)
    out = pd.DataFrame(
        {t: (df[src] if src is not None else "") for t, src in mapping.items()},
        columns=YT_TARGET,
    )
    fb = _yt_stmt_date(name)
    out["Day"] = out["Day"].map(lambda v: _yt_normalize_day(v, fb))

    if "Month" in df.columns:
        layout = "red label (assinatura)"
    elif "Day" in df.columns:
        layout = "asset raw (por canal)"
    else:
        layout = "ads/AdSense (período = data do arquivo)"
    return out, f"✅ {len(out):,} linhas · {layout}"


def _yt_fill_required(df):
    """Preenche campos que o Reprtoir exige não-vazios (senão o parser quebra):
    - Adjustment Type vazio -> 'None';
    - Asset Type vazio -> busca por Asset ID nos demais relatórios, senão 'Sound Recording'.
    (As linhas do formato ads não trazem essas colunas.)"""
    aid = df["Asset ID"].astype(str).str.strip()
    at = df["Asset Type"].astype(str).str.strip()
    have = at != ""
    lookup = dict(zip(aid[have], at[have]))
    empty_at = ~have
    if empty_at.any():
        df.loc[empty_at, "Asset Type"] = (
            aid[empty_at].map(lookup).fillna("Sound Recording").values
        )
    adj = df["Adjustment Type"].astype(str).str.strip()
    df.loc[adj == "", "Adjustment Type"] = "None"
    return df


@st.cache_data(show_spinner=False)
def _yt_consolidate(files):
    """files: tupla de (nome, bytes). Retorna (df_consolidado | None, infos)."""
    frames, infos = [], []
    for name, raw in files:
        try:
            df, info = _yt_read_one(name, raw)
        except Exception as e:
            df, info = None, f"❌ erro ao ler: {e}"
        infos.append((name, info))
        if df is not None:
            frames.append(df)
    if not frames:
        return None, infos
    consolidated = _yt_fill_required(pd.concat(frames, ignore_index=True))
    return consolidated, infos


def _yt_total(df):
    return float(pd.to_numeric(df["Partner Revenue"], errors="coerce").fillna(0).sum())


def _yt_debit_us(df):
    """Aplica débito de 30% no Partner Revenue das linhas Country=US (Decimal exato)."""
    out = df.copy()
    mask = out["Country"].astype(str).str.strip().str.upper().eq("US")

    def deb(s):
        s = str(s).strip()
        if s == "":
            return s
        try:
            d = Decimal(s)
        except Exception:
            return s
        nd = d * Decimal("0.7")            # debitar 30% == manter 70%
        return s if nd == d else format(nd, "f")

    out.loc[mask, "Partner Revenue"] = out.loc[mask, "Partner Revenue"].map(deb)
    return out, int(mask.sum())


def _yt_to_csv_bytes(df):
    """CSV no padrão aceito pelo Reprtoir: UTF-8, aspas mínimas (RFC 4180)."""
    return df.to_csv(index=False, lineterminator="\n").encode("utf-8")


def render_youtube():
    st.caption("Consolida os relatórios do YouTube (vários formatos) no template asset raw e calcula o débito de 30% (US).")

    uploaded_files = st.file_uploader(
        "Faça o upload dos relatórios do YouTube (.csv)",
        type=["csv"],
        accept_multiple_files=True,
        key="yt_files",
    )

    if not uploaded_files:
        st.info(
            "📤 Suba os relatórios do YouTube (.csv). Aceita os diferentes formatos que o YouTube "
            "envia — **asset raw** (por canal), **red label** (assinatura, com a linha 'Asset Summary') "
            "e **ads/AdSense** (sem coluna de período). São detectados automaticamente, unificados no "
            "template asset raw (período `AAAAMMDD`) e os campos exigidos pelo Reprtoir são preenchidos."
        )
        return

    files = tuple((f.name, f.getvalue()) for f in uploaded_files)
    consolidated, infos = _yt_consolidate(files)

    with st.expander(f"📋 Arquivos lidos ({len(infos)})", expanded=False):
        for nome, info in infos:
            st.markdown(f"**{nome}** — {info}")

    if consolidated is None:
        st.error("❌ Nenhum relatório válido foi reconhecido.")
        return

    total = _yt_total(consolidated)
    st.success(
        f"✅ Consolidado gerado: {len(consolidated):,} linhas · "
        f"{len(consolidated.columns)} colunas (template asset raw, período AAAAMMDD)."
    )

    col1, col2 = st.columns(2)
    with col1:
        st.metric("💰 Partner Revenue total", _yt_fmt_money(total))
    with col2:
        st.metric("🧾 Linhas consolidadas", f"{len(consolidated):,}")

    with st.expander("👁️ Visualizar consolidado (100 primeiras linhas)", expanded=False):
        st.dataframe(consolidated.head(100), use_container_width=True)

    st.download_button(
        label="📥 Baixar consolidado (sem débito)",
        data=_yt_to_csv_bytes(consolidated),
        file_name="YouTube_consolidado_asset_raw_template.csv",
        mime="text/csv",
        use_container_width=True,
        key="yt_dl_consolidado",
    )

    st.divider()
    if st.button("📉 Calcular débito de 30% (US)", type="primary", use_container_width=True):
        st.session_state["yt_debit_done"] = True

    if st.session_state.get("yt_debit_done"):
        debited, n_us = _yt_debit_us(consolidated)
        total_deb = _yt_total(debited)
        debito = total - total_deb

        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("💰 Total original", _yt_fmt_money(total))
        with c2:
            st.metric("📉 Débito US (30%)", _yt_fmt_money(debito))
        with c3:
            st.metric("✅ Total debitado", _yt_fmt_money(total_deb), delta=round(-debito, 2))

        st.caption(f"Linhas com Country = US: {n_us:,} (as de receita 0 permanecem inalteradas)")

        st.download_button(
            label="📥 Baixar consolidado DEBITADO (US -30%)",
            data=_yt_to_csv_bytes(debited),
            file_name="YouTube_consolidado_asset_raw_template_DEBITADO.csv",
            mime="text/csv",
            use_container_width=True,
            key="yt_dl_debitado",
        )


# ============================================================================
# TEMPLATE: WARNER CHAPPELL
# ============================================================================

# Ajuste do campo `royalty_period`: os statements trazem o trimestre como
# AAAAMM+AAAAMM (ex.: 202601202603 = 1T/2026). Tratamos o trimestre pelo
# ÚLTIMO mês, então mantemos apenas os 6 últimos dígitos (-> 202603).
#
# IMPORTANTE: o arquivo é processado como TEXTO CRU (round-trip latin-1),
# alterando somente o primeiro campo de cada linha. Os valores decimais de
# alta precisão (ex.: 0.00094950021848810393...) e tudo mais são preservados
# byte a byte — pandas reparsearia esses números e perderia precisão.


def _wc_fix_periods(raw):
    """Recebe os bytes do CSV e devolve (bytes_processados, mapeamentos, n_avisos).
    `mapeamentos`: dict {(periodo_original, periodo_novo): contagem}."""
    text = raw.decode("latin-1")
    lines = text.split("\n")            # LF -> split/join é lossless
    out = [lines[0]]                    # cabeçalho inalterado
    changes, issues = {}, 0
    for line in lines[1:]:
        if line.strip() == "":
            out.append(line)           # preserva eventuais linhas em branco
            continue
        comma = line.find(",")
        period = line[:comma] if comma != -1 else ""
        if comma != -1 and period.isdigit() and len(period) >= 6:
            new = period[-6:]          # mantém apenas o último mês do trimestre
            out.append(new + line[comma:])
            changes[(period, new)] = changes.get((period, new), 0) + 1
        else:
            out.append(line)           # campo inicial inesperado -> mantém intacto
            issues += 1
    return "\n".join(out).encode("latin-1"), changes, issues


def render_warner():
    st.caption(
        "Ajusta a coluna `royalty_period` do statement da Warner Chappell: mantém apenas o "
        "último mês do trimestre (ex.: `202601202603` → `202603`). Só a data muda — os valores "
        "decimais são preservados exatamente."
    )

    uploaded = st.file_uploader("Upload do statement (.csv)", type=["csv"], key="wc_file")
    if not uploaded:
        st.info("📤 Suba o statement da Warner Chappell (.csv).")
        return

    raw = uploaded.getvalue()
    processed, changes, issues = _wc_fix_periods(raw)

    total_alteradas = sum(changes.values())
    st.success(f"✅ {total_alteradas:,} linha(s) ajustada(s).")

    if changes:
        df_map = pd.DataFrame(
            [{"royalty_period (original)": o, "royalty_period (novo)": n, "Linhas": c}
             for (o, n), c in sorted(changes.items())]
        )
        st.dataframe(df_map, use_container_width=True, hide_index=True)

    if issues:
        st.warning(
            f"⚠️ {issues} linha(s) não reconhecida(s) (campo inicial não numérico) — "
            "mantidas como estavam."
        )

    with st.expander("👁️ Visualizar (100 primeiras linhas)", expanded=False):
        preview = pd.read_csv(
            io.BytesIO(processed), dtype=str, keep_default_na=False,
            encoding="latin-1", nrows=100,
        )
        st.dataframe(preview, use_container_width=True)

    base = uploaded.name.rsplit(".", 1)[0]
    st.download_button(
        label="📥 Baixar statement processado",
        data=processed,
        file_name=f"{base}_processado.csv",
        mime="text/csv",
        use_container_width=True,
        key="wc_dl",
    )


# ============================================================================
# ROTEAMENTO POR TEMPLATE
# ============================================================================

st.divider()

if template == "Nikita Digital":
    render_nikita()
elif template == "Backoffice":
    render_backoffice()
elif template == "YouTube (Consolidação)":
    render_youtube()
elif template == "Warner Chappell":
    render_warner()
